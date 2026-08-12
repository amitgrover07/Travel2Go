import openpyxl

wb = openpyxl.load_workbook('PM Framework Calculators.xlsx', data_only=False)

for name in wb.sheetnames:
    if name == 'Start Here':
        continue
    sheet = wb[name]
    print(f"=== Sheet: {name} ===")
    
    # Let's inspect the first 15 rows and 10 columns
    rows = list(sheet.iter_rows(max_row=20, max_col=12, values_only=False))
    
    # We want to see cell values and cell formulas
    for r_idx, row in enumerate(rows):
        row_str = []
        has_val = False
        for c_idx, cell in enumerate(row):
            val = cell.value
            if val is not None:
                has_val = True
                # Format cell representation
                coord = cell.coordinate
                if isinstance(val, str) and val.startswith('='):
                    row_str.append(f"{coord}: {val} (Formula)")
                else:
                    row_str.append(f"{coord}: {val}")
        if has_val:
            print(f"Row {r_idx+1}: " + " | ".join(row_str))
    print("\n")
