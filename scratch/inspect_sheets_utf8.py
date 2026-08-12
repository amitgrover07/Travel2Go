import openpyxl

wb = openpyxl.load_workbook('PM Framework Calculators.xlsx', data_only=False)

with open('scratch/sheet_details.txt', 'w', encoding='utf-8') as f:
    for name in wb.sheetnames:
        sheet = wb[name]
        f.write(f"=== Sheet: {name} ===\n")
        
        # Let's inspect the first 40 rows and 15 columns
        rows = list(sheet.iter_rows(max_row=40, max_col=15, values_only=False))
        
        # We want to see cell values and cell formulas
        for r_idx, row in enumerate(rows):
            row_str = []
            has_val = False
            for c_idx, cell in enumerate(row):
                val = cell.value
                if val is not None:
                    has_val = True
                    # Format cell representation
                    coord = cell.coordinate
                    if isinstance(val, str) and val.startswith('='):
                        row_str.append(f"{coord}: {val} (Formula)")
                    else:
                        row_str.append(f"{coord}: {val}")
            if has_val:
                f.write(f"Row {r_idx+1}: " + " | ".join(row_str) + "\n")
        f.write("\n\n")
print("Finished writing to scratch/sheet_details.txt")
