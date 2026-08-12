import openpyxl
import json

wb = openpyxl.load_workbook('PM Framework Calculators.xlsx', data_only=True)
all_data = {}

for name in wb.sheetnames:
    if name == 'Start Here':
        continue
    sheet = wb[name]
    rows = list(sheet.iter_rows(values_only=True))
    
    # We want to extract sheet title, descriptions and data
    sheet_data = {
        'title': rows[0][0] if len(rows) > 0 else '',
        'subtitle': rows[1][0] if len(rows) > 1 else '',
        'description': rows[2][0] if len(rows) > 2 else '',
        'rows': []
    }
    
    # Let's write a custom parser for each sheet's structure to capture it cleanly
    raw_rows = []
    for r in rows:
        # filter out empty rows
        if any(cell is not None for cell in r):
            raw_rows.append(list(r))
            
    sheet_data['raw_rows'] = raw_rows
    all_data[name] = sheet_data

with open('scratch/excel_data.json', 'w', encoding='utf-8') as f:
    json.dump(all_data, f, indent=2)

print("Finished dumping Excel data to scratch/excel_data.json")
